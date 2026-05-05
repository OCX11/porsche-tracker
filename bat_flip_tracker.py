#!/usr/bin/env python3
"""
BaT Flip Tracker v3
- Brave Search API with exact VIN query
- Fetches actual listing pages for price + seller
- Builds xlsx with profit analysis
"""

import sqlite3, time, re, json, random
from datetime import datetime
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH   = Path("/Users/claw/porsche-tracker/data/inventory.db")
OUTPUT    = Path("/Users/claw/Desktop/bat_flip_tracker.xlsx")
LOG_PATH  = Path("/Users/claw/porsche-tracker/data/bat_flip_results.json")
BRAVE_KEY = "***REDACTED***"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

# Sources to skip (not resale listings)
SKIP_DOMAINS = ["bringatrailer.com", "carsandbids.com", "pcarmarket.com",
                "nhtsa.gov", "vpic.nhtsa", "carfax.com", "vehiclehistory",
                "nicb.org", "vincheck", "faxvin", "vindecoderz"]

def bat_fee(hammer): return min(int(hammer * 0.05), 7500)

def label_source(url):
    if not url: return "Unknown"
    u = url.lower()
    for d, name in {
        "autotrader.com":"AutoTrader","cars.com":"Cars.com",
        "dupontregistry.com":"DuPont Registry","ebay.com":"eBay",
        "cargurus.com":"CarGurus","classic.com":"Classic.com",
        "hemmings.com":"Hemmings","pcarmarket.com":"pcarmarket",
        "rennlist.com":"Rennlist","classiccars.com":"ClassicCars",
        "motorcar.com":"Motorcar","porsche.com":"Porsche CPO",
        "dealerrater.com":"DealerRater","capitalone.com":"Capital One",
    }.items():
        if d in u: return name
    return "Dealer/Other"

def is_skip(url):
    u = url.lower()
    return any(s in u for s in SKIP_DOMAINS)

def brave_search(vin, sold_date):
    """Search Brave for exact VIN. Returns list of result dicts."""
    url = "https://api.search.brave.com/res/v1/web/search"
    params = {
        "q": f'"{vin}"',
        "count": 10,
        "search_lang": "en",
        "country": "us",
        "freshness": "py1",  # past year
    }
    headers = {
        "Accept": "application/json",
        "Accept-Encoding": "gzip",
        "X-Subscription-Token": BRAVE_KEY,
    }
    try:
        r = requests.get(url, params=params, headers=headers, timeout=12)
        if r.status_code != 200:
            print(f"  [Brave {r.status_code}]", end="")
            return []
        data = r.json()
        results = data.get("web", {}).get("results", [])
        return results
    except Exception as e:
        print(f"  [Brave err: {e}]", end="")
        return []

def extract_price_from_page(url):
    """Fetch a listing page and extract asking price + seller name."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)
        if r.status_code != 200: return None, None
        soup = BeautifulSoup(r.text, "html.parser")
        text = soup.get_text(" ", strip=True)

        # Price: look for $NNN,NNN pattern, ignore very small/large numbers
        prices = []
        for m in re.finditer(r'\$\s*([\d,]+)', text):
            n = int(m.group(1).replace(",",""))
            if 10000 < n < 2000000:
                prices.append(n)
        price = prices[0] if prices else None

        # Seller: try og:site_name, title, or domain
        seller = None
        og = soup.find("meta", property="og:site_name")
        if og: seller = og.get("content","").strip()
        if not seller:
            title = soup.find("title")
            if title: seller = title.get_text(strip=True)[:60]

        return price, seller
    except:
        return None, None

def find_listing(car):
    vin       = car["vin"]
    sold_date = car["sold_date"]
    orig_url  = car.get("listing_url","")

    results = brave_search(vin, sold_date)
    time.sleep(random.uniform(0.8, 1.5))

    # Filter results
    candidates = []
    for r in results:
        url  = r.get("url","")
        desc = r.get("description","") + " " + r.get("title","")
        if is_skip(url): continue
        if orig_url and orig_url in url: continue
        vin_in_url  = vin.lower() in url.lower()
        vin_in_desc = vin.upper() in desc.upper()
        if vin_in_url or vin_in_desc:
            candidates.append({
                "url": url,
                "title": r.get("title",""),
                "description": desc,
                "vin_in_url": vin_in_url,
                "age": r.get("age",""),
            })

    if not candidates:
        return {"status":"Unknown","confidence":"N/A","current_price":None,
                "current_source":None,"current_url":None,"seller":None}

    # Pick best: VIN in URL > VIN in description
    candidates.sort(key=lambda x: x["vin_in_url"], reverse=True)
    best = candidates[0]

    # Try to extract price from the page
    price, seller_raw = extract_price_from_page(best["url"])
    time.sleep(random.uniform(0.5, 1.0))

    # Also try price from snippet
    if not price:
        m = re.search(r'\$([\d,]+)', best["description"])
        if m:
            n = int(m.group(1).replace(",",""))
            if 10000 < n < 2000000:
                price = n

    conf = "HIGH" if best["vin_in_url"] else "MEDIUM"
    return {
        "status":           "Listed",
        "confidence":       conf,
        "current_price":    price,
        "current_source":   label_source(best["url"]),
        "current_url":      best["url"],
        "seller":           seller_raw,
    }

# ── Phase 1 ───────────────────────────────────────────────────────────────────
def fetch_comps():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = con.execute("""
        SELECT year, make, model, trim, mileage, sold_price, sold_date,
               vin, listing_url, transmission, color, generation
        FROM sold_comps
        WHERE source = 'Bring a Trailer'
          AND sold_date >= date('now', '-90 days')
          AND vin IS NOT NULL AND vin != ''
          AND sold_price IS NOT NULL
        ORDER BY sold_date DESC
    """).fetchall()
    con.close()
    print(f"[Phase 1] {len(rows)} BaT comps with VIN + sold price")
    return [dict(r) for r in rows]

# ── Phase 3: XLSX ─────────────────────────────────────────────────────────────
COLS = [
    ("VIN",16),("Year",6),("Model / Trim",28),("Mileage",10),
    ("BaT Sold Date",14),("Hammer $",12),("Buyer Fee $",11),("Total Cost $",13),
    ("Current Ask $",14),("Gross Profit $",14),("Gross Margin %",14),
    ("Source",16),("Seller / Dealer",24),("Confidence",12),("Status",11),
    ("Current URL",44),("BaT Listing",44),
]
HDR_FILL = PatternFill("solid", start_color="1F2937")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DAT_FONT = Font(name="Arial", size=10)
AMB_FILL = PatternFill("solid", start_color="FFF3CD")
UNK_FILL = PatternFill("solid", start_color="F3F4F6")
THIN     = Side(style="thin", color="E5E7EB")
BDR      = Border(bottom=THIN, right=THIN)

def build_xlsx(rows):
    wb = Workbook(); ws = wb.active
    ws.title = "BaT Flip Tracker"; ws.freeze_panes = "A2"
    for col,(name,width) in enumerate(COLS,1):
        c = ws.cell(row=1,column=col,value=name)
        c.font,c.fill = HDR_FONT,HDR_FILL
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    today = datetime.today().date()
    for i,r in enumerate(rows,2):
        car,res = r["car"],r["result"]
        hammer  = car["sold_price"]
        fee     = bat_fee(hammer)
        cost    = hammer + fee
        ask     = res.get("current_price")
        conf    = res.get("confidence","N/A")
        status  = res.get("status","Unknown")
        gross   = (ask - cost) if ask else None
        margin  = (gross/ask) if (gross and ask) else None
        trim    = " ".join(filter(None,[car.get("model",""),car.get("trim","")])).strip()
        try: days = (today - datetime.strptime(car["sold_date"],"%Y-%m-%d").date()).days
        except: days = None
        vals = [car["vin"],car["year"],trim,car.get("mileage"),car["sold_date"],
                hammer,fee,cost,ask,gross,margin,
                res.get("current_source",""),res.get("seller",""),
                conf,status,res.get("current_url",""),car.get("listing_url","")]
        rf = UNK_FILL if status=="Unknown" else (AMB_FILL if conf=="MEDIUM" else None)
        for col,val in enumerate(vals,1):
            c = ws.cell(row=i,column=col,value=val)
            c.font=DAT_FONT; c.border=BDR
            c.alignment=Alignment(vertical="center")
            if rf: c.fill=rf
            if col in (6,7,8,9,10):
                c.number_format='$#,##0'
                c.alignment=Alignment(horizontal="right",vertical="center")
                if col==10 and gross is not None:
                    c.font=Font(name="Arial",size=10,bold=True,
                                color="166534" if gross>=0 else "991B1B")
            elif col==11:
                if margin is not None: c.value=margin
                c.number_format='0.0%'
                if margin is not None:
                    c.font=Font(name="Arial",size=10,bold=True,
                                color="166534" if margin>=0 else "991B1B")
    ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
    wb.save(OUTPUT)
    print(f"\n✅  Saved → {OUTPUT}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    comps = fetch_comps()

    # Resume from checkpoint
    results, done_vins = [], set()
    if LOG_PATH.exists():
        try:
            ex = json.loads(LOG_PATH.read_text())
            if ex:
                results   = ex
                done_vins = {r["car"]["vin"] for r in results}
                print(f"[Resume] {len(done_vins)} done, {len(comps)-len(done_vins)} remaining")
        except: pass

    remaining = [c for c in comps if c["vin"] not in done_vins]
    total     = len(comps)
    print(f"\n[Phase 2] Brave-searching {len(remaining)} VINs...\n")

    for idx, car in enumerate(remaining, len(done_vins)+1):
        trim = (car.get("trim") or car.get("model") or "")[:22]
        print(f"  [{idx:>3}/{total}] {car['year']} {trim:<22} {car['vin']}", end=" → ", flush=True)
        try:
            res = find_listing(car)
        except Exception as e:
            res = {"status":"Unknown","confidence":"N/A","current_price":None,
                   "current_source":None,"current_url":None,"seller":None}
            print(f"[ERR: {e}]", end="")

        if res["status"] == "Unknown":
            print("NOT FOUND")
        else:
            price = f"${res['current_price']:,}" if res.get("current_price") else "no price"
            print(f"{res['confidence']} | {res['current_source']} | {price}")

        results.append({"car":car,"result":res})
        if idx % 25 == 0:
            LOG_PATH.write_text(json.dumps(results,indent=2))
            print(f"  ── checkpoint {idx}/{total} ──")

    LOG_PATH.write_text(json.dumps(results,indent=2))
    print(f"\n[Phase 3] Building xlsx ({len(results)} cars)...")
    build_xlsx(results)

    found   = sum(1 for r in results if r["result"]["status"]=="Listed")
    unknown = sum(1 for r in results if r["result"]["status"]=="Unknown")
    high    = sum(1 for r in results if r["result"]["confidence"]=="HIGH")
    med     = sum(1 for r in results if r["result"]["confidence"]=="MEDIUM")
    print(f"\n{'─'*50}")
    print(f"  Total:     {len(results)}")
    print(f"  Listed:    {found}  ({found/len(results)*100:.0f}%)")
    print(f"  Unknown:   {unknown}")
    print(f"  HIGH conf: {high}")
    print(f"  MED conf:  {med}  ← amber rows")
    print(f"{'─'*50}")

if __name__ == "__main__":
    main()
